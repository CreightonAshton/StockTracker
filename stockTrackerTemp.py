from bs4 import BeautifulSoup
from urllib.request import urlopen
from datetime import datetime
from openpyxl import Workbook
from openpyxl import load_workbook

wb = load_workbook('portfolios/testIndex.xlsx')
ws = wb.active
num_stocks = 10
my_portfolio = {}

for i in range(2,num_stocks+2):
	cell_of_symbol = 'A' + str(i)
	my_stock_name = ws[cell_of_symbol].value
	content = urlopen('http://finance.yahoo.com/q?s=' + my_stock_name)
	page_holder = content.read()
	content.close()
	soup = BeautifulSoup(page_holder)
	span_id = 'yfs_l84_' + my_stock_name.lower()
	my_stock_price = float(soup.find(id=span_id).string)
	print ("Current value of ", my_stock_name, " is $", my_stock_price, ".", sep="")
	my_portfolio[my_stock_name] = my_stock_price
	cell_of_current_price = 'S' + str(i)
	ws[cell_of_current_price] = my_stock_price

current_date = "{:%B %d, %Y}".format(datetime.now())
wb.save('portfolios/testIndex_' + current_date +'.xlsx')
print("Portfolio has been saved.")