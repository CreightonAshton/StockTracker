from bs4 import BeautifulSoup
from urllib.request import urlopen
from datetime import datetime
from openpyxl import Workbook
from openpyxl import load_workbook

current_date = "{:%B %d, %Y}".format(datetime.now())
current_month = "{:%B}".format(datetime.now())
num_month = "{:%m}".format(datetime.now())
current_year = "{:%Y}".format(datetime.now())
#my_portfolio = {}

def update_stocks (wb_name):
	wb = load_workbook(wb_name)
	ws = wb.active
	
	num_stocks = 0
	x = 2
	cell_counter = 'A2'
	while ws[cell_counter].value != None:
		num_stocks += 1
		x += 1
		cell_counter = 'A' + str(x)
	
	for i in range(2,num_stocks+2):
		cell_of_symbol = 'A' + str(i)
		my_stock_name = ws[cell_of_symbol].value
		content = urlopen('http://finance.yahoo.com/q?s=' + my_stock_name)
		page_holder = content.read()
		content.close()
		soup = BeautifulSoup(page_holder)
		span_id = 'yfs_l84_' + my_stock_name.lower()
		span_value = soup.find(id=span_id)
		if span_value == None:
			print(my_stock_name + " was not a recognized stock symbol")
		else:
			my_stock_price = float(span_value.string)
			print ("Current value of ", my_stock_name, " is $", my_stock_price, ".", sep="")
			#my_portfolio[my_stock_name] = my_stock_price
			cell_of_current_price = 'S' + str(i)
			ws[cell_of_current_price] = my_stock_price
	
	ans = ""
	while ans != 'y' or ans != 'n':
		ans = str(input("Would you like to save your portfolio? (y/n) "))
		if ans == 'y':
			wb.save(num_month + ' - ' + current_month + '/My Index ' + current_date + '.xlsx')
			print("Portfolio has been saved.")
			break
		elif ans == 'n':
			break
		else:
			print("Please enter either y or n")
"""
def save_stocks(wb_name):
	wb = load_workbook(wb_name)
	wb.save(num_month + ' - ' + current_month + '/My Index ' + current_date + '.xlsx')
	print("Portfolio has been saved.")
"""
def new_portfolio():
	user_name = str(input("What is your name? "))
	num_stocks = int(input("How many stocks do you want to track? "))
	wb = Workbook()
	ws = wb.active
	ws.title = user_name + "'s Portfolio"
	ws['A1'] = "Symbol"
	ws['B1'] = "Quantity"
	ws['C1'] = "Price"
	ws['D1'] = "Total*"
	ws['E1'] = "Plus Fees"
	ws['F1'] = "Jan."
	ws['G1'] = "Feb."
	ws['H1'] = "Mar."
	ws['I1'] = "Apr."
	ws['J1'] = "May"
	ws['K1'] = "Jun."
	ws['L1'] = "Jul."
	ws['M1'] = "Aug."
	ws['N1'] = "Sep."
	ws['O1'] = "Oct."
	ws['P1'] = "Nov."
	ws['Q1'] = "Dec."
	
	ws['S1'] = "Current"
	
	ws['U1'] = "Net Gain/(Loss)*"
	ws['V1'] = "Net % Gain/Loss*"
	
	ws['X1'] = "Net Gain/(Loss)"
	ws['Y1'] = "Net % Gain/Loss"
	
	
	for i in range(2, num_stocks + 2):
		my_stock_name = str(input("Stock Name: ")).upper()
		cell_of_symbol = 'A' + str(i)
		ws[cell_of_symbol] = my_stock_name
		quant = int(input("How many shares of " + my_stock_name + "? "))
		cell_of_quant = 'B' + str(i)
		ws[cell_of_quant] = quant
		orig_price = float(input("Initial Stock Price: "))
		cell_of_orig_price = 'C' + str(i)
		ws[cell_of_orig_price] = orig_price
		cell_of_subtot = 'D' + str(i)
		ws[cell_of_subtot] = orig_price * quant
		fees = float(input("Commissions/Fees: "))
		cell_of_tot = 'E' + str(i)
		ws[cell_of_tot] = (orig_price * quant) + fees
		
		content = urlopen('http://finance.yahoo.com/q?s=' + my_stock_name)
		page_holder = content.read()
		content.close()
		soup = BeautifulSoup(page_holder)
		span_id = 'yfs_l84_' + my_stock_name.lower()
		span_value = soup.find(id=span_id)
		if span_value == None:
			print(my_stock_name + " was not a recognized stock symbol")
		else:
			my_stock_price = float(span_value.string)
			print ("Current value of ", my_stock_name, " is $", my_stock_price, ".", sep="")
			cell_of_current_price = 'S' + str(i)
			ws[cell_of_current_price] = my_stock_price
	
	temp_cell_name = 'A' + str(num_stocks + 3)
	ws[temp_cell_name] = "TOTAL"
	temp_cell_name = 'A' + str(num_stocks + 5)
	ws[temp_cell_name] = "Net Gain/(Loss)*"
	temp_cell_name = 'A' + str(num_stocks + 6)
	ws[temp_cell_name] = "Net % Gain/Loss*"
	temp_cell_name = 'A' + str(num_stocks + 8)
	ws[temp_cell_name] = "Net Gain/(Loss)"
	temp_cell_name = 'A' + str(num_stocks + 9)
	ws[temp_cell_name] = "Net % Gain/Loss"
	
	ans = ""
	while ans != 'y' or ans != 'n':
		ans = str(input("Would you like to save your portfolio? (y/n) "))
		if ans == 'y':
			wb.save(num_month + ' - ' + current_month + '/' + user_name + " Portfolio - " + current_date + '.xlsx')
			print("Portfolio has been saved.")
			break
		elif ans == 'n':
			break
		else:
			print("Please enter either y or n")

	
def startup():	
	new_old = ""
	while new_old != 'new' or new_old != 'open':
		new_old = str(input("New file: new    Open existing file: open \n"))
		if new_old == "new":
			new_portfolio()
			break
		elif new_old == "open":
			wb_name = str(input("What is the name of your portfolio? "))
			file_path = wb_name + '.xlsx'
			update_stocks(file_path)
			break
		else:
			print ("Please enter either new or open.")

startup()			
#update_stocks('My Index ' + current_year + '.xlsx')
#save_stocks('My Index ' + current_year + '.xlsx')

