from openpyxl import Workbook
from openpyxl import load_workbook
"""
wb = Workbook()
ws = wb.active
ws1 = wb.create_sheet()
ws2 = wb.create_sheet(0)
ws.title = "New Title"
ws3 = wb["New Title"]
ws4 = wb.get_sheet_by_name("New Title")
print (ws is ws3 is ws4)
print (wb.get_sheet_names())
for sheet in wb:
	print(sheet.title)
	
c = ws['A4']
ws['A4'] = 4
c = ws.cell('A4')
d = ws.cell(row=4, column=2)

cell_range = ws['A1':'C2']
print (cell_range)
print(ws.range('A1:C2'))
for row in ws.range('A1:C2'):
	for cell in row:
		print (cell)

ws = wb.active
ws['C9'] = "hello world"
print(ws.rows)
print(ws.columns)

c.value = '12%'
print (c.value)
import datetime
d.value = datetime.datetime.now()
print (d.value)
c.value = '31.50'
print(c.value)

wb.save('portfolios/testing.xlsx')
"""

wb2 = load_workbook('portfolios/testIndex.xlsx')
print (wb2.get_sheet_names())